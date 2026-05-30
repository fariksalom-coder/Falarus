#!/usr/bin/env python3
"""Apply official docx answer keys to patent_questions_new_version.xlsx (variants 12–16)."""
from __future__ import annotations

import json
import re
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "scripts"))

from extract_patent_answers import DEFAULT_DOCX, extract_answers  # noqa: E402
from import_patent_exam import parse_audio_question_numbers  # noqa: E402

XLSX = Path.home() / "Downloads" / "patent_questions_new_version.xlsx"
BACKUP = XLSX.with_name("patent_questions_new_version.before_official_key.xlsx")
VARIANTS = [12, 13, 14, 15, 16]


def row_question_numbers(row_id: str, row_type: str) -> list[int]:
    if row_type == "audioDouble":
        return parse_audio_question_numbers(row_id)
    _, _, chunk = row_id.split("_")
    return [int(chunk)]


def apply_row(
    row_id: str,
    row_type: str,
    answer_type: str,
    qmap: dict[int, int | str],
    questions_json: str | None,
    sub_json: str | None,
    correctanswers: str | None,
) -> tuple[str | None, str | None, str | None]:
    numbers = row_question_numbers(row_id, row_type)
    if row_type == "audioDouble":
        subs = json.loads(sub_json or "[]")
        if len(subs) != len(numbers):
            raise ValueError(f"{row_id}: sub-question count mismatch")
        for num, sub in zip(numbers, subs):
            ans = qmap[num]
            if not isinstance(ans, int):
                raise ValueError(f"{row_id} Q{num}: expected choice index, got {ans!r}")
            sub["correct"] = ans
        return None, json.dumps(subs, ensure_ascii=False), correctanswers

    num = numbers[0]
    ans = qmap[num]
    if answer_type == "written":
        if not isinstance(ans, str):
            raise ValueError(f"{row_id} Q{num}: expected written answer, got {ans!r}")
        return questions_json, sub_json, ans

    payload = json.loads(questions_json or "{}")
    if not isinstance(ans, int):
        raise ValueError(f"{row_id} Q{num}: expected choice index, got {ans!r}")
    payload["correct"] = ans
    return json.dumps(payload, ensure_ascii=False), sub_json, correctanswers


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing workbook: {XLSX}")
    if not DEFAULT_DOCX.exists():
        raise SystemExit(f"Missing docx: {DEFAULT_DOCX}")

    answers = extract_answers(DEFAULT_DOCX, VARIANTS)
    for v in VARIANTS:
        missing = [q for q in range(1, 23) if q not in answers[v]]
        if missing:
            raise SystemExit(f"V{v} missing keys: {missing}")

    if not BACKUP.exists():
        shutil.copy2(XLSX, BACKUP)
        print(f"backup -> {BACKUP}")

    wb = load_workbook(XLSX)
    ws = wb.active
    header = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col_id = header["id"]
    col_type = header["type"]
    col_answer = header["answerType"]
    col_q = header["QuestionsJson"]
    col_sub = header["subQuestionsJson"]
    col_written = header["Correctanswers"]

    updated = 0
    for r in range(2, ws.max_row + 1):
        row_id = ws.cell(r, col_id).value
        if not row_id or not re.match(r"^P_(12|13|14|15|16)_", str(row_id)):
            continue
        vnum = int(str(row_id).split("_")[1])
        row_type = ws.cell(r, col_type).value
        answer_type = ws.cell(r, col_answer).value
        qj = ws.cell(r, col_q).value
        sj = ws.cell(r, col_sub).value
        ca = ws.cell(r, col_written).value
        new_qj, new_sj, new_ca = apply_row(
            str(row_id), str(row_type), str(answer_type), answers[vnum], qj, sj, ca
        )
        changed = False
        if new_qj != qj:
            ws.cell(r, col_q).value = new_qj
            changed = True
        if new_sj != sj:
            ws.cell(r, col_sub).value = new_sj
            changed = True
        if new_ca != ca:
            ws.cell(r, col_written).value = new_ca
            changed = True
        if changed:
            updated += 1
            print(f"updated {row_id}")

    wb.save(XLSX)
    print(f"Saved {XLSX} ({updated} rows changed)")


if __name__ == "__main__":
    main()
